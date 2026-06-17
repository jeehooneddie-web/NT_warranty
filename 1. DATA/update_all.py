"""
update_all.py
─────────────
OneDrive Excel → 집계 → HTML embed → git push 원클릭 자동화
실행: python update_all.py  또는  업데이트.bat 더블클릭
"""
import pandas as pd
import json, re, os, shutil, subprocess, sys
from openpyxl import load_workbook

BASE      = 'D:/NM-Dev/warranty-dashboard'
DATA_DIR  = f'{BASE}/1. DATA'
HTML_PATH = f'{BASE}/dashboard-app/preview/index.html'
ONEDRIVE  = 'C:/Users/user/OneDrive - 내쇼날모터스/보증팀 원드라이브/1. Warranty Claim RAW'

RAW_SRC      = f'{ONEDRIVE}/2026 data RAW_Claim.xlsx'
RAW_TMP      = f'{DATA_DIR}/raw_claim_tmp.xlsx'
DEFECT_SRC   = f'{ONEDRIVE}/디펙트코드 리스트.xlsx'
DEFECT_TMP   = f'{DATA_DIR}/defect_list_tmp.xlsx'
TC_SRC       = f'{ONEDRIVE}/RecallTcRptRawData.xlsx'
TC_TMP       = f'{DATA_DIR}/tc_raw_tmp.xlsx'
TC_ARCHIVE   = f'C:/Users/user/OneDrive - 내쇼날모터스/보증팀 원드라이브/TC_누적_RAW.xlsx'
EXCLUDE_SRC  = f'{ONEDRIVE}/RECALL TC 작업불가 리스트.xlsx'
EXCLUDE_TMP  = f'{DATA_DIR}/tc_exclude_tmp.xlsx'
KR_REJECT_SRC = f'{ONEDRIVE}/KR_REJECT_LIST.xlsx'
KR_REJECT_TMP = f'{DATA_DIR}/kr_reject_tmp.xlsx'

def step(msg): print(f'\n[▶] {msg}')
def ok(msg):   print(f'    ✓ {msg}')
def err(msg):  print(f'    ✗ {msg}'); sys.exit(1)

SKIP_COPY = '--skip-copy' in sys.argv

# ── 1. 파일 복사 ──────────────────────────────────────────
if SKIP_COPY:
    step('파일 복사 건너뜀 (기존 임시파일 사용)')
    _kr_reject_ok = os.path.exists(KR_REJECT_TMP)
    ok('기존 임시파일로 진행')
else:
    step('OneDrive 파일 복사 중...')
    for src, dst, name in [
        (RAW_SRC,    RAW_TMP,    'RAW 클레임'),
        (DEFECT_SRC, DEFECT_TMP, '디펙트코드 리스트'),
        (TC_SRC,     TC_TMP,     'TC 미실시 RAW'),
        (EXCLUDE_SRC, EXCLUDE_TMP, '작업불가 리스트'),
    ]:
        try:
            shutil.copy2(src, dst)
            ok(f'{name} 복사 완료')
        except PermissionError:
            err(f'{name} 파일이 Excel에서 열려 있습니다. 닫고 다시 실행하세요.')
        except FileNotFoundError:
            err(f'{name} 파일을 찾을 수 없습니다: {src}')

    # KR_REJECT_LIST는 선택 파일 — 없거나 열려있어도 계속 진행
    _kr_reject_ok = False
    try:
        shutil.copy2(KR_REJECT_SRC, KR_REJECT_TMP)
        ok('KR Reject 리스트 복사 완료')
        _kr_reject_ok = True
    except PermissionError:
        ok('KR Reject 리스트가 Excel에서 열려 있음 (무시)')
    except FileNotFoundError:
        ok('KR Reject 리스트 없음 (무시)')

# ── 2. RAW 클레임 읽기 ────────────────────────────────────
step('RAW 클레임 집계 중...')
# header 행 자동 감지: B열(index 1)이 '청구/취소' 또는 '청구'를 포함하는 행을 헤더로 사용
_raw = pd.read_excel(RAW_TMP, sheet_name=2, header=None, engine='openpyxl', nrows=5)
_header_row = 0
for i in range(5):
    val = str(_raw.iloc[i, 1])
    if '청구' in val or '취소' in val or 'Claim' in val:
        _header_row = i
        break
df = pd.read_excel(RAW_TMP, sheet_name=2, header=_header_row, engine='openpyxl')
cols = df.columns.tolist()
ok(f'헤더 행: {_header_row}행  /  총 컬럼: {len(cols)}개  /  B열: {cols[1]}')
status_col = cols[1]; city_col = cols[5]; code_col = cols[12]
amount_col = cols[22]; task_col = cols[8]; person_col = cols[23]
k_col = cols[10]  # K열: Warranty Stage
# month/type은 수식열(AC/AE) 대신 원본열(N열 날짜, K/M열)로 직접 계산
df['_month'] = pd.to_datetime(df[cols[13]], errors='coerce').dt.strftime('%y-%m')
month_col = '_month'

def _compute_type(k_val, m_val):
    k = str(k_val).strip() if pd.notna(k_val) else ''
    m = str(m_val).strip() if pd.notna(m_val) else ''
    if 'goodwill' in k.lower(): return 'Goodwill'
    if k.lower() == 'warranty plus': return 'WP'
    if 'LA' in m: return 'LOCAL TC'
    try:
        m_num = float(m.replace(',',''))
        if m_num < 100_000_000: return 'TC/RECALL'
        if m_num > 8_700_000_000 and not m.startswith('99'): return 'BSI'
        return 'Warranty'
    except Exception:
        return 'Warranty'

df['_type'] = df.apply(lambda r: _compute_type(r[k_col], r[code_col]), axis=1)
type_col = '_type'
ok(f'month/type 직접 계산 완료 (수식열 비의존)')

# 지점 코드 → 지점명 매핑 (숫자코드 또는 AS_ 형식 모두 처리)
BRANCH_MAP = {'26999':'전주','40699':'군산','41967':'목포','419668':'서산','44835':'평택'}

TYPES = ['BSI','Warranty','TC/RECALL','WP','LOCAL TC','Goodwill']
df_c = df[df[status_col].str.contains('청구', na=False)].copy()
df_c = df_c[df_c[type_col].isin(TYPES)].copy()
df_c[city_col] = df_c[city_col].astype(str).str.replace('AS_','', regex=False)
df_c[city_col] = df_c[city_col].replace(BRANCH_MAP)
df_c[code_col]  = df_c[code_col].astype(str).str.strip()
df_c[amount_col] = pd.to_numeric(df_c[amount_col], errors='coerce').fillna(0)
# M열이 '99'로 시작하는 경우 BSI → Warranty 재분류 (99코드는 BSI 아님)
mask_99 = (df_c[type_col] == 'BSI') & (df_c[code_col].str.startswith('99'))
df_c.loc[mask_99, type_col] = 'Warranty'
ok(f'청구 행 수: {len(df_c):,}  /  99코드 재분류: {mask_99.sum():,}건 (BSI→Warranty)')

# ── 3. BRANCH_DATA 집계 ───────────────────────────────────
step('BRANCH_DATA 집계 중...')
df_branch_count = df_c[df_c[task_col] == 1].copy()
branch_cnt = df_branch_count.groupby([city_col, month_col, type_col]).agg(count=(amount_col,'count')).reset_index()
branch_tot = df_c.groupby([city_col, month_col, type_col]).agg(total=(amount_col,'sum')).reset_index()
branch_agg = branch_cnt.merge(branch_tot, on=[city_col, month_col, type_col], how='outer').fillna(0)
branch_agg['count'] = branch_agg['count'].astype(int)

branch_data = {}
for _, r in branch_agg.iterrows():
    city = r[city_col]; mo = r[month_col]; tp = r[type_col]
    branch_data.setdefault(city, {}).setdefault(mo, {})[tp] = {
        'count': int(r['count']), 'total': int(r['total'])
    }

branch_js = 'const BRANCH_DATA = ' + json.dumps(branch_data, ensure_ascii=False, separators=(',',':')) + ';'
ok(f'지점 수: {len(branch_data)}')

# ── 4. 디펙트코드 리스트 읽기 ─────────────────────────────
step('디펙트코드 리스트 읽기 중...')
dl = pd.read_excel(DEFECT_TMP, sheet_name=0, header=0, engine='openpyxl')
dl_cols = dl.columns.tolist()
desc_map = {str(r[dl_cols[0]]).strip(): str(r[dl_cols[1]]).strip() for _, r in dl.iterrows()}
ok(f'코드 수: {len(desc_map):,}')

# ── 5. DEFECT_RAW 집계 ────────────────────────────────────
step('DEFECT_RAW 집계 중...')
agg = df_c.groupby([city_col, month_col, type_col, code_col]).agg(
    count=(amount_col,'count'), amount=(amount_col,'sum')
).reset_index()

matched_desc = {}
unmatched_set = set()
for code in agg[code_col].unique():
    if code in desc_map:
        matched_desc[code] = desc_map[code]
    else:
        unmatched_set.add(code)

defect_raw = [
    [r[city_col], r[month_col], r[type_col], r[code_col], int(r['count']), int(r['amount'])]
    for _, r in agg.iterrows()
]

um_agg = agg[agg[code_col].isin(unmatched_set)].groupby(code_col).agg(
    count=('count','sum'), amount=('amount','sum')
).reset_index().sort_values('count', ascending=False)
unmatched = [{'code':r[code_col],'count':int(r['count']),'amount':int(r['amount'])} for _,r in um_agg.iterrows()]

ok(f'집계 행: {len(defect_raw):,}  /  미매칭: {len(unmatched)}개')

# ── 5-1. DEFECT_PRICE_STAT 집계 (T/U/V열: 부품/공임/외주) ──────────
step('DEFECT_PRICE_STAT 집계 중...')
parts_col  = cols[19]  # T열: 부품
labor_col  = cols[20]  # U열: 공임
outsrc_col = cols[21]  # V열: 외주
for c in [parts_col, labor_col, outsrc_col]:
    df_c[c] = pd.to_numeric(df_c[c], errors='coerce').fillna(0)

price_agg = df_c.groupby([code_col, city_col, type_col]).agg(
    cnt       = (parts_col,  'count'),
    p_max     = (parts_col,  'max'),
    p_min     = (parts_col,  'min'),
    p_sum     = (parts_col,  'sum'),
    l_max     = (labor_col,  'max'),
    l_min     = (labor_col,  'min'),
    l_sum     = (labor_col,  'sum'),
    o_max     = (outsrc_col, 'max'),
    o_min     = (outsrc_col, 'min'),
    o_sum     = (outsrc_col, 'sum'),
    t_max     = (amount_col, 'max'),
    t_min     = (amount_col, 'min'),
    t_sum     = (amount_col, 'sum'),
).reset_index()

price_stat = {}
for _, r in price_agg.iterrows():
    code = str(r[code_col])
    price_stat.setdefault(code, []).append([
        str(r[city_col]), str(r[type_col]), int(r['cnt']),
        int(r['p_max']), int(r['p_min']), int(r['p_sum']),
        int(r['l_max']), int(r['l_min']), int(r['l_sum']),
        int(r['o_max']), int(r['o_min']), int(r['o_sum']),
        int(r['t_max']), int(r['t_min']), int(r['t_sum']),
    ])
ok(f'DEFECT_PRICE_STAT 코드 수: {len(price_stat):,}')

# ── 6. 개인별 청구 현황 집계 (Task=1, 금액=클레임금액 cols[22]) ──────────
step('개인별 청구 현황 집계 중...')
task_col   = cols[8]
person_col = cols[23]
# 금액: 보증매출(BRANCH_DATA)과 동일하게 cols[22] 사용
p_amt_col  = cols[22]

# 청구 + 취소 모두 포함
df_person_base = df[df[status_col].str.contains('청구|취소', na=False)].copy()
df_person_base = df_person_base[df_person_base[type_col].isin(TYPES)].copy()
df_person_base[city_col] = df_person_base[city_col].str.replace('AS_','', regex=False)
df_person_base[p_amt_col] = pd.to_numeric(df_person_base[p_amt_col], errors='coerce').fillna(0)

# 개인별: Task=1 필터
df_task1 = df_person_base[df_person_base[task_col] == 1].copy()

person_agg = df_task1.groupby([person_col, month_col, type_col, status_col]).agg(
    count=(p_amt_col, 'count'), amount=(p_amt_col, 'sum')
).reset_index()

person_raw = [
    [r[person_col], r[month_col], r[type_col], r[status_col], int(r['count']), int(r['amount'])]
    for _, r in person_agg.iterrows()
]

# 전체합계: Task 필터 없이 전체 집계 (보증매출과 동일 기준)
total_agg = df_person_base.groupby([month_col, type_col, status_col]).agg(
    count=(p_amt_col, 'count'), amount=(p_amt_col, 'sum')
).reset_index()
for _, r in total_agg.iterrows():
    person_raw.append(['전체합계', r[month_col], r[type_col], r[status_col], int(r['count']), int(r['amount'])])

ok(f'개인별 집계 행: {len(person_raw)}  /  고유 ID: {df_task1[person_col].nunique()}명')

# ── 6-2. 일별 현황 집계 (당월 기준, Task=1, 청구) ──────────────────────
step('일별 현황 집계 중...')
import datetime as _dt
_cur_month = _dt.date.today().strftime('%y-%m')
df_daily = df[df[status_col].str.contains('청구', na=False)].copy()
df_daily = df_daily[df_daily[type_col].isin(TYPES)].copy()
df_daily = df_daily[df_daily[month_col] == _cur_month].copy()
df_daily['_date'] = pd.to_datetime(df_daily[cols[13]], errors='coerce').dt.strftime('%y-%m-%d')
df_daily[amount_col] = pd.to_numeric(df_daily[amount_col], errors='coerce').fillna(0)
df_daily_t1 = df_daily[df_daily[task_col] == 1].copy()
daily_agg = df_daily_t1.groupby([person_col, '_date', type_col]).agg(
    count=(amount_col, 'count'), amount=(amount_col, 'sum')
).reset_index()
daily_raw = [
    [r[person_col], r['_date'], r[type_col], int(r['count']), int(r['amount'])]
    for _, r in daily_agg.iterrows()
]
ok(f'일별 집계: {len(daily_raw):,}행 (당월 {_cur_month})')

# ── 7. TC 미실시율 집계 ────────────────────────────────────
step('TC 미실시율 집계 중...')

# 작업불가 리스트 읽기
exclude_df = pd.read_excel(EXCLUDE_TMP, sheet_name=0, header=None, engine='openpyxl')
exclude_codes = set(
    str(v).strip() for v in exclude_df.iloc[:, 0].dropna()
    if str(v).strip() and str(v).strip() != 'nan'
)
exclude_list_raw = []
for _, row in exclude_df.iterrows():
    code   = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
    branch = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
    reason = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''
    if code and code != 'nan':
        exclude_list_raw.append({'code': code, 'branch': branch, 'reason': reason})
ok(f'작업불가 코드: {len(exclude_codes)}개')

# ── 누적 아카이브 병합 (당해년도 신규 + 이전 연도 누적) ──
tc_new = pd.read_excel(TC_TMP, header=0, engine='openpyxl')
if os.path.exists(TC_ARCHIVE):
    tc_old = pd.read_excel(TC_ARCHIVE, header=0, engine='openpyxl')
    tc_df = pd.concat([tc_old, tc_new], ignore_index=True)
else:
    tc_df = tc_new
_dedup_keys = [tc_df.columns[1], tc_df.columns[4], tc_df.columns[6], tc_df.columns[11]]
tc_df = tc_df.drop_duplicates(subset=_dedup_keys, keep='last')
tc_df.to_excel(TC_ARCHIVE, index=False, engine='openpyxl')
ok(f'TC 누적 아카이브: {len(tc_df):,}행 저장')

tc_cols = tc_df.columns.tolist()
tc_campaign = tc_cols[1]; tc_car = tc_cols[4]; tc_dealer = tc_cols[6]
tc_ro = tc_cols[11]; tc_result = tc_cols[9]; tc_date = tc_cols[5]

tc_dedup = tc_df.drop_duplicates(subset=[tc_campaign, tc_car, tc_dealer, tc_ro])
tc_clean = tc_dedup[tc_dedup[tc_ro].notna() & (tc_dedup[tc_ro].astype(str).str.strip() != '')].copy()

# 작업불가 캠페인코드 제외
before = len(tc_clean)
tc_clean = tc_clean[~tc_clean[tc_campaign].astype(str).str.strip().isin(exclude_codes)].copy()
ok(f'작업불가 제외: {before - len(tc_clean):,}행 제거 → 잔여 {len(tc_clean):,}행')

# 발행일자(D열) 기준 방문일 대비 5년 이내만 포함
tc_issue = tc_cols[3]
tc_clean['_visit'] = pd.to_datetime(tc_clean[tc_date], errors='coerce')
tc_clean['_issue'] = pd.to_datetime(tc_clean[tc_issue], errors='coerce')
before = len(tc_clean)
tc_clean = tc_clean[(tc_clean['_visit'] - tc_clean['_issue']).dt.days / 365.25 <= 5].copy()
ok(f'5년 초과 제외: {before - len(tc_clean):,}행 제거 → 잔여 {len(tc_clean):,}행')
tc_clean['branch'] = tc_clean[tc_dealer].astype(str).str.replace('AS_', '', regex=False)
tc_clean['month']  = tc_clean['_visit'].dt.strftime('%y-%m')
tc_clean['is_N']   = tc_clean[tc_result].astype(str).str.strip() == 'Y'  # Y=미실시

tc_agg = tc_clean.groupby(['branch', 'month']).agg(
    total=('is_N', 'count'), n_count=('is_N', 'sum')
).reset_index()

tc_data = {}
for _, r in tc_agg.iterrows():
    tc_data.setdefault(r['branch'], {})[r['month']] = {
        'total': int(r['total']), 'n_count': int(r['n_count'])
    }
ok(f'TC 집계 완료: {len(tc_clean)}행 → {sum(len(v) for v in tc_data.values())}개월치')

# TC_TOP_DATA: 캠페인별 미실시 집계 (지점/월/캠페인코드/캠페인명)
tc_name = tc_cols[2]
tc_top_agg = tc_clean.groupby(['branch', 'month', tc_campaign, tc_name]).agg(
    total=('is_N', 'count'), n_count=('is_N', 'sum')
).reset_index()
tc_top_data = [
    [r['branch'], r['month'], str(r[tc_campaign]), str(r[tc_name]), int(r['total']), int(r['n_count'])]
    for _, r in tc_top_agg.iterrows()
]
ok(f'TC TOP 집계: {len(tc_top_data)}행')

# 당월 발행 TC (발행월 == 입고월)
same_month_mask = tc_clean['_issue'].dt.to_period('M') == tc_clean['_visit'].dt.to_period('M')
tc_same = tc_clean[same_month_mask].copy()
tc_same_agg = tc_same.groupby(['branch', 'month', tc_campaign, tc_name]).agg(
    total=('is_N', 'count'), n_count=('is_N', 'sum')
).reset_index()
tc_same_data = [
    [r['branch'], r['month'], str(r[tc_campaign]), str(r[tc_name]), int(r['total']), int(r['n_count'])]
    for _, r in tc_same_agg.iterrows()
]
ok(f'당월발행 TC: {len(tc_same_data)}건')

# ── 8. CLAIM_DATA 집계 (불승인/보완요청/보완완료, Task=1) ─
step('CLAIM_DATA 집계 중...')
df_claim_all = df[df[cols[6]].isin(['불승인', '보완요청', '보완완료'])].copy()
df_claim_all[cols[22]] = pd.to_numeric(df_claim_all[cols[22]], errors='coerce').fillna(0)

# 클레임번호별 전체 금액 합산 (Task 무관)
claim_amount_total = df_claim_all.groupby(cols[7])[cols[22]].sum().to_dict()

# 클레임현황 시트에서 Credit 수신일자 추출 (승인여부='승인대기' 행 기준)
credit_map = {}
try:
    df_kh2 = pd.read_excel(RAW_TMP, sheet_name='클레임현황', header=0, engine='openpyxl')
    kc2 = df_kh2.columns.tolist()
    kh2_no     = kc2[2]   # 클레임번호
    kh2_credit = kc2[13]  # Credit 수신일자
    kh2_appr   = kc2[22]  # 승인여부
    df_reject_rows = df_kh2[df_kh2[kh2_appr] == '승인대기']
    for _, row in df_reject_rows.iterrows():
        cno = str(row[kh2_no])
        cdt = row[kh2_credit]
        if pd.notna(cdt) and str(cdt) not in ('nan', 'NaT'):
            credit_map[cno] = str(cdt)[:10]
    ok(f'Credit 수신일자 매핑: {len(credit_map)}건')
except Exception as e:
    ok(f'Credit 수신일자 로드 실패 (무시): {e}')

# Task=1 행만 표시 (금액은 클레임번호 단위 합산값으로 대체)
df_claim = df_claim_all[df_claim_all[cols[8]] == 1].copy()
claim_raw = []
for _, r in df_claim.iterrows():
    date_val = str(r[cols[13]])[:10] if pd.notna(r[cols[13]]) and str(r[cols[13]]) not in ('nan','NaT') else ''
    claim_no = str(r[cols[7]])
    claim_raw.append([
        claim_no,                                             # [0] H: 클레임번호
        str(r[cols[5]]).replace('AS_',''),                    # [1] F: 지점
        str(r[cols[6]]),                                      # [2] G: 클레임상태
        r[type_col],                                          # [3] ClaimType (_compute_type 결과 — 다른 카테고리와 동일 기준)
        str(r[cols[11]]),                                     # [4] L: Claim Type (DMS 원본)
        str(r[cols[12]]),                                     # [5] M: Defect Code
        date_val,                                             # [6] N: 날짜
        int(claim_amount_total.get(claim_no, 0)),             # [7] W: 클레임금액 (전체합산)
        str(r[cols[23]]),                                     # [8] X: 담당확인자
        str(r[cols[18]]) if pd.notna(r[cols[18]]) else '',   # [9] S: 차대번호
        credit_map.get(claim_no, ''),                         # [10] Credit 수신일자
    ])
ok(f'CLAIM_DATA: {len(claim_raw):,}행 (불승인/보완요청/보완완료, Task=1)')

# ── 9. WHOLESALE_DATA 집계 (클레임현황 시트) ──────────────────
step('WHOLESALE_DATA 집계 중...')
wholesale_data = {}
try:
    df_kh = pd.read_excel(RAW_TMP, sheet_name='클레임현황', header=0, engine='openpyxl')
    kc = df_kh.columns.tolist()
    # 열 인덱스: 지점=0, J=9, M=12, O=14, R=17
    kh_J = kc[9];  kh_M = kc[12]; kh_O = kc[14]; kh_R = kc[17]
    # 헤더 신구 호환 (BAT 실행 전은 구 헤더 S/T/U/W/Y 사용)
    kh_T = 'Claim Type'   if 'Claim Type'   in kc else 'T'
    kh_W = '승인여부'      if '승인여부'      in kc else 'W'
    kh_S = '클레임확정월'  if '클레임확정월'  in kc else kc[18]  # S열: 승인대기 기준월
    kh_U = '보증마감기준'  if '보증마감기준'  in kc else 'U'
    kh_Y = '회계마감기준'  if '회계마감기준'  in kc else 'Y'
    for col in [kh_J, kh_M, kh_O, kh_R]:
        df_kh[col] = pd.to_numeric(df_kh[col], errors='coerce').fillna(0)
    for col in [kh_T, kh_W, kh_S, kh_U, kh_Y]:
        df_kh[col] = df_kh[col].astype(str).str.strip()

    kh_branch = kc[0]  # 지점 컬럼 (index 0)
    df_kh[kh_branch] = df_kh[kh_branch].astype(str).str.strip().str.replace('AS_','', regex=False)
    WS_BRANCHES = ['전주', '군산', '목포', '서산', '평택']

    def _agg_claimtype(df_sub, mo_col):
        """단일 데이터프레임 → {ClaimType: {month: {charge/approve/pending}}}"""
        df_v = df_sub[df_sub[mo_col].str.match(r'\d{2}-\d{2}', na=False)]
        out = {}
        for _, r in df_v.groupby([kh_T, mo_col]).agg(
            amount=(kh_M,'sum'), parts=(kh_J,'sum'), count=(kh_M,'count')
        ).reset_index().iterrows():
            ct, mo = str(r[kh_T]), str(r[mo_col])
            out.setdefault(ct, {}).setdefault(mo, {})['charge'] = {
                'amount': int(r['amount']), 'parts': int(r['parts']), 'count': int(r['count'])
            }
        for _, r in df_v.groupby([kh_T, mo_col]).agg(
            amount=(kh_R,'sum'), parts=(kh_O,'sum'), count=(kh_R,'count')
        ).reset_index().iterrows():
            ct, mo = str(r[kh_T]), str(r[mo_col])
            out.setdefault(ct, {}).setdefault(mo, {})['approve'] = {
                'amount': int(r['amount']), 'parts': int(r['parts']), 'count': int(r['count'])
            }
        df_pend = df_sub[
            (df_sub[kh_W] == '승인대기') &
            (df_sub[kh_S].str.match(r'\d{2}-\d{2}', na=False))
        ]
        for _, r in df_pend.groupby([kh_T, kh_S]).agg(
            amount=(kh_M,'sum'), parts=(kh_J,'sum'), count=(kh_M,'count')
        ).reset_index().iterrows():
            ct, mo = str(r[kh_T]), str(r[kh_S])
            out.setdefault(ct, {}).setdefault(mo, {})['pending'] = {
                'amount': int(r['amount']), 'parts': int(r['parts']), 'count': int(r['count'])
            }
        return out

    def _agg_basis(df_base, mo_col):
        """전체 + 지점별 집계 → {지점: {ClaimType: {month: {...}}}}"""
        result = {'전체': _agg_claimtype(df_base, mo_col)}
        for br in WS_BRANCHES:
            df_br = df_base[df_base[kh_branch] == br]
            if len(df_br) > 0:
                result[br] = _agg_claimtype(df_br, mo_col)
        return result

    wholesale_data = {
        '보증마감': _agg_basis(df_kh, kh_U),
        '회계마감': _agg_basis(df_kh, kh_Y),
    }
    ok(f'보증마감 지점 수: {len(wholesale_data["보증마감"])}  /  회계마감: {len(wholesale_data["회계마감"])}  /  총 행: {len(df_kh):,}')
except Exception as _e:
    ok(f'클레임현황 시트 없음 또는 오류 → WHOLESALE_DATA 빈값 ({_e})')

# ── 8-0. QR_CLAIM_DATA 집계 (Claim 상세_전체, 2026년 수리일 기준) ──
step('QR_CLAIM_DATA 집계 중...')
try:
    df_cl = pd.read_excel(RAW_TMP, sheet_name='Claim 상세_전체', header=1, engine='openpyxl')
    cc = df_cl.columns.tolist()
    cc_claim=cc[7]; cc_defect=cc[12]; cc_vin=cc[18]
    cc_branch=cc[5]; cc_cdate=cc[13]  # N열: 클레임생성일자
    cc_status=cc[1]  # B열: 청구/취소
    cc_stage=cc[10]  # K열: 보증 Stage
    cc_parts=cc[19]  # T열: 부품금액

    def _qr_type(stage, defect):
        s = str(stage).strip() if pd.notna(stage) else ''
        d = str(defect).strip() if pd.notna(defect) else ''
        if 'goodwill' in s.lower(): return 'Goodwill'
        if 'warranty plus' in s.lower(): return 'WP'
        if d.upper().startswith('LA'): return 'LOCAL TC'
        try:
            num = float(d.replace(',',''))
            if num < 100_000_000: return 'TC/RECALL'
            if num > 8_700_000_000 and not d.startswith('99'): return 'BSI'
        except Exception: pass
        return 'Warranty'
    df_cl['_cdate'] = pd.to_datetime(df_cl[cc_cdate], errors='coerce')
    df_cl_2026 = df_cl[
        df_cl[cc_claim].notna() &
        df_cl[cc_claim].astype(str).str.startswith('WC') &
        (df_cl['_cdate'].dt.year >= 2026) &
        df_cl[cc_status].astype(str).str.contains('청구', na=False)
    ].copy()
    qr_seen = set()
    qr_claim_rows = []
    for _, r in df_cl_2026.iterrows():
        cn = str(r[cc_claim]).strip()[2:]
        defect = r[cc_defect]
        if pd.isna(defect): continue
        key = (cn, str(defect).strip())
        if key in qr_seen: continue
        qr_seen.add(key)
        vin = str(r[cc_vin]).strip() if pd.notna(r[cc_vin]) else ''
        vin7 = vin[-7:] if len(vin) >= 7 else vin
        branch = str(r[cc_branch]).replace('AS_','').strip()
        try:
            cdate_str = pd.to_datetime(r[cc_cdate]).strftime('%Y-%m-%d') if pd.notna(r[cc_cdate]) else ''
        except Exception:
            cdate_str = ''
        qr_type = _qr_type(r[cc_stage], defect)
        try:
            parts_amt = int(float(str(r[cc_parts]).replace(',', ''))) if pd.notna(r[cc_parts]) else 0
        except Exception:
            parts_amt = 0
        qr_claim_rows.append([cn, str(defect).strip(), vin7, branch, cdate_str, qr_type, parts_amt])
    ok(f'QR_CLAIM_DATA: {len(qr_claim_rows):,}건 (2026년)')
except Exception as _e:
    qr_claim_rows = []
    ok(f'QR_CLAIM_DATA 생성 실패 (무시): {_e}')

# ── 8-1. KR_REJECT_DATA 집계 ──────────────────────────────
step('KR_REJECT_DATA 집계 중...')
kr_reject_rows = []
try:
    if _kr_reject_ok:
        KR_BRANCH_MAP = {'26999':'전주','40699':'군산','41967':'목포',
                         '41968':'서산','419668':'서산','44835':'평택'}
        PERSON_NAMES = {
            'bna157':'이지훈','bna255':'최금환','bna358':'김민철','bna377':'강청호',
            'bna472':'기준혁','bna557':'유정현','nt230403':'최희원','nt250204':'고서진',
            'nt250601':'신기섭','nt250702':'고형민','nt250911':'김선기','nt250917':'한현구',
            'nt251001':'권승리','nt251104':'오정훈',
        }
        # RAW claim data에서 WC번호(WC제거) → 담당자 lookup
        # 단독키(Ext.No) + 복합키(Ext.No|VIN) 두 가지 저장 — 복합키 우선 조회
        claim_person_map  = {}   # Ext.No 단독
        claim_person_map2 = {}   # Ext.No + VIN 복합
        for _, row in df.iterrows():
            cno    = str(row[cols[7]]).upper().replace('WC', '').strip()
            vin    = str(row[cols[18]]).strip()
            person = str(row[cols[23]]).strip()
            if cno and cno != 'nan' and person and person != 'nan':
                claim_person_map[cno] = person
                claim_person_map2[cno + '|' + vin] = person

        kr_df = pd.read_excel(KR_REJECT_TMP, sheet_name='KR REJECT LIST',
                              header=None, engine='openpyxl')
        for _, row in kr_df.iloc[1:].iterrows():   # 헤더행(0) 스킵
            # 중간 헤더 반복 행 스킵 (Ext.No가 숫자가 아닌 행)
            if not pd.notna(row.iloc[7]) or not str(row.iloc[7]).replace('.','',1).isdigit():
                continue
            claim_no  = str(row.iloc[1]).strip()  if pd.notna(row.iloc[1])  else ''
            ext_no    = str(int(float(row.iloc[7])))
            vin_val   = str(row.iloc[2]).strip()  if pd.notna(row.iloc[2])  else ''
            hst       = str(int(float(row.iloc[8]))) if pd.notna(row.iloc[8]) else ''
            ref_date  = str(row.iloc[4]).split(' ')[0] if pd.notna(row.iloc[4]) else ''
            rtype     = str(row.iloc[9]).strip()  if pd.notna(row.iloc[9])  else ''
            reason    = str(row.iloc[15]).strip() if pd.notna(row.iloc[15]) else ''
            month     = str(row.iloc[26]).strip() if pd.notna(row.iloc[26]) and str(row.iloc[26]) != 'nan' else ''
            confirmed = str(row.iloc[28]).strip() if pd.notna(row.iloc[28]) and str(row.iloc[28]) != 'nan' else ''
            branch    = KR_BRANCH_MAP.get(hst, hst)
            # 복합키(Ext.No+VIN) 우선, 없으면 단독키(Ext.No) fallback
            person_id = claim_person_map2.get(ext_no + '|' + vin_val, '') or claim_person_map.get(ext_no, '')
            person    = PERSON_NAMES.get(person_id, person_id)
            kr_reject_rows.append([claim_no, branch, ref_date, rtype, reason, person, confirmed, month])
        ok(f'KR_REJECT_DATA: {len(kr_reject_rows):,}건')
    else:
        ok('KR_REJECT_DATA: 파일 없음, 빈 배열 사용')
except Exception as _e:
    kr_reject_rows = []
    ok(f'KR_REJECT_DATA 생성 실패 (무시): {_e}')

# AB열(담당자) 자동 기입 — 데이터 집계와 분리하여 실패해도 kr_reject_rows 유지
if kr_reject_rows:
    try:
        wb = load_workbook(KR_REJECT_SRC)
        ws = wb['KR REJECT LIST']
        ws.cell(row=1, column=28).value = '담당자'
        for i, r in enumerate(kr_reject_rows):
            ws.cell(row=i + 2, column=28).value = r[5]
        wb.save(KR_REJECT_SRC)
        ok('AB열(담당자) 자동 기입 완료')
    except Exception as _e2:
        ok(f'AB열 자동 기입 실패 (데이터는 정상): {_e2}')

# ── 8-2. PERSON_DEFECT_RAW 집계 (청구 전체 × 담당자 × 타입 × 지점 × DefectCode × 월) ──
# 포맷: [personId, month, claimType, branch, defectCode, count, amount]
step('PERSON_DEFECT_RAW 집계 중...')
try:
    df_pd = df_c[[person_col, month_col, type_col, city_col, code_col, amount_col]].copy()
    df_pd = df_pd[df_pd[person_col].notna() & (df_pd[person_col].astype(str).str.strip() != '')]
    df_pd[person_col] = df_pd[person_col].astype(str).str.strip()
    df_pd[code_col]   = df_pd[code_col].astype(str).str.strip()
    pd_agg = df_pd.groupby([person_col, month_col, type_col, city_col, code_col]).agg(
        count=(amount_col, 'count'),
        amount=(amount_col, 'sum')
    ).reset_index()
    person_defect_raw = [
        [str(r[person_col]), str(r[month_col]), str(r[type_col]), str(r[city_col]), str(r[code_col]), int(r['count']), int(r['amount'])]
        for _, r in pd_agg.iterrows()
    ]
    ok(f'PERSON_DEFECT_RAW: {len(person_defect_raw):,}행')
except Exception as _e:
    person_defect_raw = []
    ok(f'PERSON_DEFECT_RAW 생성 실패 (무시): {_e}')

# ── 8-1. JS 문자열 생성 ───────────────────────────────────
step('JS 데이터 문자열 생성 중...')
desc_js     = 'const DEFECT_DESC='       + json.dumps(matched_desc, ensure_ascii=False, separators=(',',':')) + ';'
raw_js      = 'const DEFECT_RAW='        + json.dumps(defect_raw,   ensure_ascii=False, separators=(',',':')) + ';'
um_js       = 'const DEFECT_UNMATCHED='  + json.dumps(unmatched,    ensure_ascii=False, separators=(',',':')) + ';'
price_js    = 'const DEFECT_PRICE_STAT=' + json.dumps(price_stat,   ensure_ascii=False, separators=(',',':')) + ';'
person_js   = 'const PERSON_DATA='    + json.dumps(person_raw,   ensure_ascii=False, separators=(',',':')) + ';'
daily_js    = 'const PERSON_DAILY='   + json.dumps(daily_raw,    ensure_ascii=False, separators=(',',':')) + ';'
tc_js       = 'const TC_DATA='        + json.dumps(tc_data,      ensure_ascii=False, separators=(',',':')) + ';'
tc_top_js   = 'const TC_TOP_DATA='   + json.dumps(tc_top_data,  ensure_ascii=False, separators=(',',':')) + ';'
tc_same_js  = 'const TC_SAME_MONTH_DATA=' + json.dumps(tc_same_data, ensure_ascii=False, separators=(',',':')) + ';'
claim_js      = 'const CLAIM_DATA='      + json.dumps(claim_raw,      ensure_ascii=False, separators=(',',':')) + ';'
wholesale_js  = 'const WHOLESALE_DATA=' + json.dumps(wholesale_data, ensure_ascii=False, separators=(',',':')) + ';'
qr_claim_js        = 'const QR_CLAIM_DATA='       + json.dumps(qr_claim_rows,      ensure_ascii=False, separators=(',',':')) + ';'
person_defect_js   = 'const PERSON_DEFECT_RAW='  + json.dumps(person_defect_raw,  ensure_ascii=False, separators=(',',':')) + ';'
kr_reject_js       = 'const KR_REJECT_DATA='     + json.dumps(kr_reject_rows,     ensure_ascii=False, separators=(',',':')) + ';'
ok('완료')

# ── 7. HTML embed ─────────────────────────────────────────
step('HTML에 데이터 삽입 중...')
with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

html = re.sub(r'const BRANCH_DATA\s*=\s*\{.*?\};',  branch_js,  html, flags=re.DOTALL)
html = re.sub(r'const DEFECT_DESC=\{.*?\};',          desc_js,    html, flags=re.DOTALL)
html = re.sub(r'const DEFECT_RAW=\[.*?\];',           raw_js,     html, flags=re.DOTALL)
html = re.sub(r'const DEFECT_UNMATCHED=\[.*?\];',     um_js,      html, flags=re.DOTALL)
html = re.sub(r'const DEFECT_PRICE_STAT=\{.*?\};',    price_js,   html, flags=re.DOTALL)
html = re.sub(r'const PERSON_DATA=\[.*?\];',         person_js,  html, flags=re.DOTALL)
html = re.sub(r'const PERSON_DAILY=\[.*?\];',        daily_js,   html, flags=re.DOTALL)
html = re.sub(r'const TC_DATA=\{.*?\};',             tc_js,      html, flags=re.DOTALL)
html = re.sub(r'const TC_TOP_DATA=\[.*?\];',         tc_top_js,  html, flags=re.DOTALL)
html = re.sub(r'const TC_SAME_MONTH_DATA=\[.*?\];', tc_same_js, html, flags=re.DOTALL)
html = re.sub(r'const CLAIM_DATA=\[.*?\];',          claim_js,      html, flags=re.DOTALL)
html = re.sub(r'const WHOLESALE_DATA=\{.*?\};',     wholesale_js,  html, flags=re.DOTALL)
html = re.sub(r'const QR_CLAIM_DATA=\[.*?\];',      qr_claim_js,        html, flags=re.DOTALL)
html = re.sub(r'const PERSON_DEFECT_RAW=\[.*?\];', person_defect_js,   html, flags=re.DOTALL)
html = re.sub(r'const KR_REJECT_DATA=\[.*?\];',   kr_reject_js,        html, flags=re.DOTALL)

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html)
ok(f'HTML 크기: {os.path.getsize(HTML_PATH)/1024:.1f}KB')

# ── 7-1. 외부 페이지 재빌드 (4개 뷰 고정 공개) ───────────────
step('외부 페이지(external/index.html) 재빌드 중...')
EXT_HTML_PATH = f'{BASE}/dashboard-app/preview/external/index.html'
try:
    from datetime import datetime as _dt
    ext_date = _dt.now().strftime('%Y-%m-%d %H:%M')

    # ① index.html에서 HTML 마커 추출
    with open(HTML_PATH, 'r', encoding='utf-8') as f:
        idx_content = f.read()

    ext_views_html = {}
    for vid in ['view-sales', 'view-top-defect', 'view-tc', 'view-stage23']:
        m = re.search(f'<!-- EXT:{vid}:START -->(.*?)<!-- EXT:{vid}:END -->', idx_content, re.DOTALL)
        ext_views_html[vid] = m.group(1) if m else ''
        ok(f'{vid} HTML 추출: {len(ext_views_html[vid])}chars')

    # ② index.html에서 JS 마커 추출
    ext_js = {}
    for jid in ['theme', 'sales', 'defect', 'tc', 'stage23']:
        m = re.search(f'// EXT_JS:{jid}:START(.*?)// EXT_JS:{jid}:END', idx_content, re.DOTALL)
        ext_js[jid] = m.group(1) if m else ''
        ok(f'JS:{jid} 추출: {len(ext_js[jid])}chars')

    # ③ external/index.html 읽기
    with open(EXT_HTML_PATH, 'r', encoding='utf-8') as f:
        ext_html = f.read()

    # ④ HTML 인젝트 (lambda로 백슬래시 이스케이프 문제 우회)
    for vid, html_content in ext_views_html.items():
        repl = f'<!-- EXT_INJECT:{vid}:START -->{html_content}<!-- EXT_INJECT:{vid}:END -->'
        ext_html = re.sub(
            f'<!-- EXT_INJECT:{vid}:START -->.*?<!-- EXT_INJECT:{vid}:END -->',
            lambda m, r=repl: r, ext_html, flags=re.DOTALL
        )

    # ⑤ JS 인젝트 (lambda로 백슬래시 이스케이프 문제 우회)
    for jid, js_content in ext_js.items():
        repl = f'// EXT_JS_INJECT:{jid}:START{js_content}// EXT_JS_INJECT:{jid}:END'
        ext_html = re.sub(
            f'// EXT_JS_INJECT:{jid}:START.*?// EXT_JS_INJECT:{jid}:END',
            lambda m, r=repl: r, ext_html, flags=re.DOTALL
        )

    # ⑥ 데이터 embed (4개 뷰 항상 전체)
    subs = [
        (r'const EXT_BRANCH_DATA=[\[{].*?[\]}];',
         'const EXT_BRANCH_DATA=' + json.dumps(branch_data,   ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_DEFECT_RAW=\[.*?\];',
         'const EXT_DEFECT_RAW='  + json.dumps(defect_raw,    ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_DEFECT_DESC=\{.*?\};',
         'const EXT_DEFECT_DESC=' + json.dumps(matched_desc,  ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_DEFECT_UNMATCHED=\[.*?\];',
         'const EXT_DEFECT_UNMATCHED=' + json.dumps(unmatched, ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_TC_DATA=[\[{].*?[\]}];',
         'const EXT_TC_DATA='     + json.dumps(tc_data,       ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_TC_TOP_DATA=\[.*?\];',
         'const EXT_TC_TOP_DATA=' + json.dumps(tc_top_data,   ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_TC_SAME_DATA=\[.*?\];',
         'const EXT_TC_SAME_DATA='+ json.dumps(tc_same_data,  ensure_ascii=False, separators=(',',':')) + ';'),
        (r'const EXT_WHOLESALE_DATA=[\[{].*?[\]}];',
         'const EXT_WHOLESALE_DATA=' + json.dumps(wholesale_data, ensure_ascii=False, separators=(',',':')) + ';'),
        (r"const EXT_UPDATE_DATE='.*?';",
         f"const EXT_UPDATE_DATE='{ext_date}';"),
    ]
    for pattern, replacement in subs:
        ext_html = re.sub(pattern, replacement, ext_html, flags=re.DOTALL)

    with open(EXT_HTML_PATH, 'w', encoding='utf-8') as f:
        f.write(ext_html)
    ok(f'외부 페이지 재빌드 완료: {os.path.getsize(EXT_HTML_PATH)/1024:.1f}KB')
except Exception as _e:
    import traceback; traceback.print_exc()
    print(f'    ⚠ 외부 페이지 재빌드 실패 (무시): {_e}')

# ── 8. git commit & push ──────────────────────────────────
step('git commit & push 중...')
os.chdir(BASE)
subprocess.run(['git', 'add', 'dashboard-app/preview/index.html', 'dashboard-app/preview/external/index.html'], check=True)

from datetime import datetime
today = datetime.now().strftime('%Y-%m-%d %H:%M')

# 변경사항이 있을 때만 커밋
status = subprocess.run(['git', 'status', '--porcelain', 'dashboard-app/preview/index.html'],
                        capture_output=True, text=True).stdout.strip()
if status:
    subprocess.run(['git', 'commit', '-m', f'데이터 자동 업데이트 ({today})'], check=True)
    subprocess.run(['git', 'push'], check=True)
    ok('배포 완료!')
else:
    ok('데이터 변경 없음 — 커밋 스킵 (이미 최신 상태)')

# ── 9. Azure Static Web Apps 배포 ────────────────────────
step('Azure Static Web Apps 배포 중...')
try:
    sys.path.insert(0, DATA_DIR)
    from config_azure import SWA_DEPLOYMENT_TOKEN
    APP_DIR = f'{BASE}/dashboard-app/preview'
    result = subprocess.run(
        ['swa.cmd', 'deploy', APP_DIR,
         '--deployment-token', SWA_DEPLOYMENT_TOKEN,
         '--env', 'production'],
        capture_output=True, text=True, errors='replace'
    )
    if result.returncode == 0:
        ok('SWA 배포 완료! → https://warranty.nationalmotors.co.kr')
    else:
        print(f'    ⚠ SWA 배포 실패: {result.stderr}')
except ImportError as e:
    print(f'    ⚠ config_azure.py 없음: {e}')
except Exception as e:
    print(f'    ⚠ SWA 배포 오류: {e}')

print('\n' + '='*50)
print('  모든 작업이 완료됐습니다.')
print('='*50 + '\n')
